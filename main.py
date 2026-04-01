import os, re, json, sys, subprocess, tempfile, threading, webbrowser
from collections import defaultdict
from pathlib import Path
from datetime import datetime

# ── Auto-instalar dependencias ────────────────────────────────────────────────
for pkg in ("pdfplumber", "openpyxl", "flask", "pymupdf", "psycopg2-binary", "sqlalchemy"):
    try:
        __import__(pkg if pkg != "flask" else "flask")
    except ImportError:
        print(f"Instalando {pkg}...")
        subprocess.run([sys.executable, "-m", "pip", "install", pkg,
                        "--break-system-packages", "-q"], check=True)

import pdfplumber, openpyxl, fitz  # fitz = pymupdf
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, request, jsonify, send_file, send_from_directory

# ── Parche para CockroachDB ─────────────────────────────────────────────────
from sqlalchemy.dialects.postgresql.psycopg2 import PGDialect_psycopg2

# Parcheamos directamente la clase del dialecto PostgreSQL
original_get_server_version = PGDialect_psycopg2._get_server_version_info

def patched_get_server_version(self, connection):
    try:
        return original_get_server_version(self, connection)
    except AssertionError:
        # Si falla el parseo de versión (CockroachDB), devolvemos una versión fija
        print("⚠️  Detectada versión de CockroachDB, usando versión fija (25.4.1)")
        return (25, 4, 1)

# Aplicamos el parche
PGDialect_psycopg2._get_server_version_info = patched_get_server_version

# ── Base de Datos ───────────────────────────────────────────────────────────
from sqlalchemy import create_engine, Column, String, Integer, Text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import NullPool

app = Flask(__name__, static_folder=".")

# ── Rutas de archivos ─────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent
PRODUCTOS_TXT = BASE_DIR / "productos.txt"
OUTPUT_XLSX   = BASE_DIR / "resumen_pedidos.xlsx"

# ── Configuración de Base de Datos ──────────────────────────────────────────
DATABASE_URL = os.getenv('DATABASE_URL', 'sqlite:///productos_local.db')

# Detectar si es CockroachDB (solo para información)
if DATABASE_URL and 'cockroachlabs.cloud' in DATABASE_URL:
    print("✅ Conectando a CockroachDB (con parche de versión)")

print(f"🔌 Conectando a BD...")

# Crear engine con configuración especial
engine = create_engine(
    DATABASE_URL,
    poolclass=NullPool,
    connect_args={
        "connect_timeout": 15,
        "application_name": "pulguitas"
    }
)

SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

class ProductoDB(Base):
    __tablename__ = 'productos'
    
    id = Column(Integer, primary_key=True)
    categoria = Column(String(50))
    modelo = Column(String(100))
    color = Column(String(50))
    talle = Column(String(10))
    texto = Column(Text, unique=True)

class CategoriaDB(Base):
    __tablename__ = 'categorias'
    
    id = Column(Integer, primary_key=True)
    nombre = Column(String(50), unique=True, nullable=False)
    color_fondo = Column(String(6), default="607D8B")
    color_texto = Column(String(6), default="FFFFFF")

# Crear tablas
Base.metadata.create_all(engine)

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE BASE DE DATOS
# ══════════════════════════════════════════════════════════════════════════════
def get_all_productos():
    """Obtiene todos los productos de la base de datos"""
    session = SessionLocal()
    productos = session.query(ProductoDB).all()
    session.close()
    return productos

def guardar_productos_desde_json(contenido_json):
    """Guarda todos los productos desde un JSON (reemplaza todo)"""
    try:
        datos = json.loads(contenido_json)
        
        session = SessionLocal()
        # Limpiar tabla
        session.query(ProductoDB).delete()
        
        # Insertar nuevos productos
        for categoria, modelos in datos.items():
            for modelo, variantes in modelos.items():
                for variante in variantes:
                    producto = ProductoDB(
                        categoria=categoria,
                        modelo=modelo,
                        color=variante.get("color", ""),
                        talle=variante.get("talle", ""),
                        texto=variante["texto"].lower()
                    )
                    session.add(producto)
        
        session.commit()
        session.close()
        return True
    except Exception as e:
        print(f"❌ Error guardando en BD: {e}")
        return False

def cargar_mapeo_desde_bd():
    """Carga el MAPA_PRODUCTOS desde la base de datos"""
    try:
        productos = get_all_productos()
        mapa_plano = {}
        for p in productos:
            mapa_plano[p.texto] = (p.categoria, p.modelo, p.color, p.talle)
        print(f"✅ Mapeo cargado desde BD: {len(mapa_plano)} entradas")
        return mapa_plano
    except Exception as e:
        print(f"❌ Error cargando desde BD: {e}")
        return {}

def exportar_mapeo_a_json():
    """Exporta todos los productos a formato JSON anidado (para frontend)"""
    productos = get_all_productos()
    
    resultado = {}
    for p in productos:
        if p.categoria not in resultado:
            resultado[p.categoria] = {}
        if p.modelo not in resultado[p.categoria]:
            resultado[p.categoria][p.modelo] = []
        resultado[p.categoria][p.modelo].append({
            "texto": p.texto,
            "color": p.color,
            "talle": p.talle
        })
    
    return resultado

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE CATEGORÍAS
# ══════════════════════════════════════════════════════════════════════════════
def get_all_categorias():
    """Obtiene todas las categorías"""
    session = SessionLocal()
    categorias = session.query(CategoriaDB).order_by(CategoriaDB.nombre).all()
    session.close()
    return categorias

def guardar_categoria(nombre, color_fondo="607D8B", color_texto="FFFFFF"):
    """Guarda una nueva categoría"""
    session = SessionLocal()
    try:
        cat = CategoriaDB(
            nombre=nombre.upper(),
            color_fondo=color_fondo,
            color_texto=color_texto
        )
        session.add(cat)
        session.commit()
        return True
    except Exception as e:
        print(f"Error guardando categoría: {e}")
        return False
    finally:
        session.close()

def eliminar_categoria(nombre):
    """Elimina una categoría"""
    session = SessionLocal()
    try:
        session.query(CategoriaDB).filter(CategoriaDB.nombre == nombre.upper()).delete()
        session.commit()
        return True
    except Exception as e:
        print(f"Error eliminando categoría: {e}")
        return False
    finally:
        session.close()

def actualizar_colores_categoria(nombre, color_fondo, color_texto):
    """Actualiza los colores de una categoría"""
    session = SessionLocal()
    try:
        cat = session.query(CategoriaDB).filter(CategoriaDB.nombre == nombre.upper()).first()
        if cat:
            cat.color_fondo = color_fondo
            cat.color_texto = color_texto
            session.commit()
            return True
        return False
    except Exception as e:
        print(f"Error actualizando colores: {e}")
        return False
    finally:
        session.close()

def get_colores_categorias():
    """Obtiene los colores de todas las categorías desde la BD"""
    session = SessionLocal()
    try:
        # Intentar obtener de la BD
        categorias = session.query(CategoriaDB).all()
        colores = {}
        for cat in categorias:
            colores[cat.nombre] = (cat.color_fondo, cat.color_texto)
        
        # Si no hay categorías, insertar las básicas
        if not colores:
            categorias_base = [
                ("VERANO", "29B6F6", "E1F5FE"),
                ("ANTIESTRES", "66BB6A", "E8F5E9"),
                ("INVIERNO", "FFA726", "FFF3E0"),
                ("DECO", "AB47BC", "F3E5F5"),
                ("ESCALERA", "78909C", "ECEFF1"),
                ("NORDICA", "26A69A", "E0F2F1"),
                ("ROPITA", "EC407A", "FCE4EC"),
                ("MANTA", "FFCA28", "FFFDE7"),
                ("DISPENSER", "9C27B0", "F3E5F5"),
            ]
            for nombre, fondo, texto in categorias_base:
                cat = CategoriaDB(nombre=nombre, color_fondo=fondo, color_texto=texto)
                session.add(cat)
            session.commit()
            
            # Recargar
            categorias = session.query(CategoriaDB).all()
            for cat in categorias:
                colores[cat.nombre] = (cat.color_fondo, cat.color_texto)
            print("✅ Categorías básicas insertadas en la BD")
        
        return colores
    except Exception as e:
        print(f"⚠️ Error cargando colores: {e}")
        return {}
    finally:
        session.close()

# Cargar el mapa al inicio desde BD
MAPA_PRODUCTOS = cargar_mapeo_desde_bd()

# Cargar colores desde la BD
CAT_COLORS = get_colores_categorias()
print(f"🎨 Colores cargados para {len(CAT_COLORS)} categorías")

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN DINÁMICA DESDE MAPEO
# ══════════════════════════════════════════════════════════════════════════════
def cargar_configuracion_desde_mapeo():
    """
    Extrae la configuración (modelos, colores, talles) desde MAPA_PRODUCTOS
    """
    modelos = set()
    colores_simples = set()
    colores_compuestos = {}
    talles = set()
    
    for texto, info in MAPA_PRODUCTOS.items():
        cat, modelo, color, talle = info
        
        # Agregar modelo (limpiar sufijos como " (solo funda)")
        modelo_limpio = modelo.replace(' (solo funda)', '').lower()
        modelos.add(modelo_limpio)
        
        # Procesar color
        if color:
            if '/' in color:
                color_lower = color.lower()
                variantes_color = [
                    color_lower,
                    color_lower.replace('/', ' '),
                    color_lower.replace('/', '')
                ]
                colores_compuestos[color_lower] = variantes_color
            else:
                colores_simples.add(color.lower())
        
        # Procesar talle
        if talle:
            talles.add(talle.upper())
    
    config = {
        "modelos": sorted(list(modelos)),
        "colores": {
            "simples": sorted(list(colores_simples)),
            "compuestos": colores_compuestos
        },
        "talles": sorted(list(talles)),
        "palabras_prohibidas": ["argentina", "boca", "river", "inter", "miami", "panda"]
    }
    
    print(f"✅ Configuración cargada: {len(config['modelos'])} modelos, {len(config['colores']['simples'])} colores simples, {len(config['colores']['compuestos'])} colores compuestos")
    return config

def generar_palabras_clave():
    """Genera palabras clave dinámicamente desde MAPA_PRODUCTOS"""
    palabras = set()
    for texto_entrada in MAPA_PRODUCTOS.keys():
        primera_palabra = texto_entrada.split()[0] if texto_entrada.split() else ""
        if len(primera_palabra) > 2:
            palabras.add(primera_palabra.lower())
    
    for info in MAPA_PRODUCTOS.values():
        modelo = info[1].lower()
        palabras_modelo = modelo.split()
        for palabra in palabras_modelo:
            if len(palabra) > 2:
                palabras.add(palabra.lower())
    
    # Agregar categorías importantes
    categorias_importantes = ["cama", "sofa", "mini", "escalera", "manta", "gatito", 
                             "nordica", "pancho", "garra", "timoteo", "remeras", "buzo", "huella"]
    for palabra in categorias_importantes:
        palabras.add(palabra)
    
    return palabras

# Cargar configuración
CONFIG = cargar_configuracion_desde_mapeo()
PALABRAS_CLAVE = generar_palabras_clave()
print(f"📋 Palabras clave generadas: {len(PALABRAS_CLAVE)}")

def normalizar_texto_sin_medidas(texto):
    """
    Normaliza un texto eliminando medidas y caracteres que pueden variar
    """
    if not texto:
        return ""
    
    import re
    
    texto = texto.lower()
    
    # Eliminar medidas
    texto = re.sub(r'\d+\s*(cm|mm|mt|m)?\s*(de\s*(alto|ancho|largo))?', ' ', texto)
    texto = re.sub(r'\b(alto|ancho|largo|cm|mm|mt)\b', ' ', texto)
    texto = re.sub(r'[\(\)\[\]\{\}]', ' ', texto)
    texto = re.sub(r'[,;:\.\-_]', ' ', texto)
    texto = re.sub(r'\s*-\s*', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto)
    texto = texto.strip()
    
    # 🔥 Después de normalizar, agregar "talla" antes de los talles sueltos
    # Pero solo si no hay "talla" ya presente
    if 'talla' not in texto:
        # Buscar talles sueltos (s, m, l, xl, xs) como palabras independientes
        texto = re.sub(r'\b(xs|s|m|l|xl)\b', r'talla \1', texto)
    
    return texto

def extraer_caracteristicas(texto):
    """
    Extrae características clave de un texto
    """
    if not texto:
        return {}
    
    texto_original = texto.lower()
    texto_sin_medidas = normalizar_texto_sin_medidas(texto)
    
    caracteristicas = {
        'modelo': None,
        'color': None,
        'talle': None,
        'lado': None,
        'tipo': 'completa'
    }
    
    # Detectar modelo
    for modelo in CONFIG['modelos']:
        if modelo in texto_sin_medidas:
            caracteristicas['modelo'] = modelo
            break
    
    # Detectar color
    for color_compuesto, variantes in CONFIG['colores']['compuestos'].items():
        for variante in variantes:
            if variante in texto_sin_medidas:
                caracteristicas['color'] = color_compuesto
                break
        if caracteristicas['color']:
            break
    
    if not caracteristicas['color']:
        for color_simple in CONFIG['colores']['simples']:
            if color_simple in texto_sin_medidas:
                caracteristicas['color'] = color_simple
                break
    
    # Detectar talle
    for t in CONFIG['talles']:
        t_lower = t.lower()
        if re.search(r'talla\s*' + t_lower, texto_original) or \
           re.search(r'talle\s*' + t_lower, texto_original) or \
           re.search(r'\b' + t_lower + r'\b', texto_original):
            caracteristicas['talle'] = t.upper()
            break
    
    # Detectar lado
    if 'derecha' in texto_sin_medidas:
        caracteristicas['lado'] = 'derecha'
    elif 'izquierda' in texto_sin_medidas:
        caracteristicas['lado'] = 'izquierda'
    
    # Detectar tipo
    if 'funda' in texto_sin_medidas and 'completa' not in texto_sin_medidas:
        caracteristicas['tipo'] = 'funda'
    elif 'solo funda' in texto_sin_medidas:
        caracteristicas['tipo'] = 'funda'
    
    return caracteristicas

# ══════════════════════════════════════════════════════════════════════════════
# CATÁLOGO
# ══════════════════════════════════════════════════════════════════════════════
TALLE_LABELS = {"S":"Talle S","M":"Talle M","L":"Talle L","XL":"Talle XL",
                "XS":"Talle XS","SM":"Talle S/M","LXL":"Talle L/XL","U":"Talle Único"}
TALLE_SIZES  = {
    "VERANO":{"S":"50x50","M":"70x70","L":"90x90"},
    "ANTIESTRES":{"M":"70x70","L":"90x90"},
    "INVIERNO":{"S":"50x50","M":"70x70","L":"90x90"},
    "DECO":{"S":"S/M","M":"M/L","L":"L/XL"},
    "ESCALERA":{"M":"30x38","L":"40x38"},
    "NORDICA":{"M":"60x60","L":"80x80","XL":"90x90"},
    "ROPITA":{"XS":"XS","SM":"S/M","LXL":"L/XL"},
    "MANTA":{"U":"70x70"},
}

def cargar_catalogo(texto):
    catalogo_dict = {}; order = []
    for linea in texto.split("\n"):
        linea = linea.strip()
        if not linea or linea.startswith("#"): continue
        partes = [p.strip() for p in linea.split("|")]
        if len(partes) < 3: continue
        cat = partes[0].upper(); modelo = partes[1]
        color  = partes[2] if len(partes) > 2 else ""
        talles = [t.strip() for t in partes[3].split(",")] if len(partes) > 3 else []
        if cat not in catalogo_dict:
            catalogo_dict[cat] = {"talle_cols": talles, "filas": []}; order.append(cat)
        else:
            for t in talles:
                if t not in catalogo_dict[cat]["talle_cols"]:
                    catalogo_dict[cat]["talle_cols"].append(t)
        catalogo_dict[cat]["filas"].append((modelo, color))
    catalogo = []
    for cat in order:
        info = catalogo_dict[cat]; tc = info["talle_cols"]
        sizes = TALLE_SIZES.get(cat, {})
        headers = ["Modelo","Color"] + [
            f"{TALLE_LABELS.get(t,t)} ({sizes[t]})" if sizes.get(t) else TALLE_LABELS.get(t,t)
            for t in tc
        ]
        while len(headers) < 5: headers.append("")
        catalogo.append({"cat":cat,"headers":headers[:5],"talle_cols":tc,"filas":info["filas"]})
    return catalogo

# ══════════════════════════════════════════════════════════════════════════════
# RESOLVER (usa el mapa cargado desde BD)
# ══════════════════════════════════════════════════════════════════════════════
def resolver(nombre):
    """
    Resuelve un nombre de producto de manera flexible buscando palabras clave
    con pesos específicos por categoría - IGNORA MEDIDAS
    """
    if not nombre:
        return None
    
    # Normalizar el texto de entrada SIN MEDIDAS
    key_original = nombre.lower().strip()
    key_sin_medidas = normalizar_texto_sin_medidas(nombre)
    
    print(f"\n🔍 Texto original: {key_original[:100]}...")
    print(f"🔍 Texto sin medidas: {key_sin_medidas[:100]}...")
    
    # Palabras que causan falsos positivos
    palabras_prohibidas = CONFIG['palabras_prohibidas']
    
    # 1. Primero intentar con coincidencia exacta del texto sin medidas
    mejor_match_exacto = None
    mejor_len_exacto = 0
    matches_exactos_con_funda = []
    matches_exactos_sin_funda = []
    
    print("\n🔎 Buscando matches exactos...")
    for p, info in MAPA_PRODUCTOS.items():
        p_sin_medidas = normalizar_texto_sin_medidas(p)
        
        # Crear una versión del texto del PDF sin "solo funda de repuesto" para comparar
        key_sin_funda = key_sin_medidas.replace("solo funda de repuesto", "").strip()
        key_sin_funda = key_sin_funda.replace("solo funda", "").strip()
        key_sin_funda = key_sin_funda.replace("funda de repuesto", "").strip()
        key_sin_funda = key_sin_funda.replace("funda", "").strip()
        
        # Limpiar espacios múltiples que puedan quedar
        key_sin_funda = re.sub(r'\s+', ' ', key_sin_funda).strip()
        
        # Verificar si el texto del mapa está en el texto del PDF (con o sin funda)
        if p_sin_medidas in key_sin_medidas or p_sin_medidas in key_sin_funda:
            # Verificar si este producto es funda o no
            es_funda_mapa = "(solo funda)" in info[1].lower() or "funda" in p.lower()
            texto_habla_de_funda = "funda" in key_sin_medidas or "repuesto" in key_sin_medidas or "solo funda" in key_sin_medidas
            
            print(f"\n  🔍 Match exacto candidato: '{p_sin_medidas}'")
            print(f"    - Producto: {info[1]} {info[3]} {info[2]}")
            print(f"    - ¿Producto es funda? {es_funda_mapa}")
            print(f"    - ¿Texto habla de funda? {texto_habla_de_funda}")
            print(f"    - Longitud base: {len(p_sin_medidas)}")
            
            # Calcular un puntaje base (longitud del match)
            puntaje_match = len(p_sin_medidas)
            
            # Si el texto habla de funda y el producto ES funda, dar PRIORIDAD MÁXIMA
            if texto_habla_de_funda and es_funda_mapa:
                puntaje_match += 1000
                print(f"    🏆 PRIORIDAD FUNDA: +1000")
            
            # Si hay consistencia (ambos funda o ambos no funda), dar bonus
            if es_funda_mapa == texto_habla_de_funda:
                puntaje_match += 100
                print(f"    ✅ CONSISTENTE: +100")
            else:
                # Inconsistencia, guardar como alternativa pero no considerar ahora
                matches_exactos_con_funda.append((puntaje_match, info))
                print(f"    ⚠️ INCONSISTENTE: guardado como alternativa")
                continue
            
            # Bonus adicional si el match es EXACTAMENTE igual (con funda incluida)
            if texto_habla_de_funda and p_sin_medidas in key_sin_medidas:
                puntaje_match += 50
                print(f"    📌 MATCH EXACTO CON FUNDA: +50")
            
            if puntaje_match > mejor_len_exacto:
                mejor_len_exacto = puntaje_match
                mejor_match_exacto = info
                print(f"    ➡️ NUEVO MEJOR MATCH (puntaje: {puntaje_match})")

    # Si encontramos un match exacto consistente, lo usamos
    if mejor_match_exacto:
        print(f"\n  ✅ Match exacto CONSISTENTE seleccionado: {mejor_match_exacto}")
        return mejor_match_exacto

    # Si no hay match consistente, pero hay matches inconsistentes
    if matches_exactos_con_funda:
        # Ordenar por puntaje (el más alto primero)
        matches_exactos_con_funda.sort(key=lambda x: -x[0])
        print(f"\n  ⚠️ Usando match exacto INCONSISTENTE como fallback: {matches_exactos_con_funda[0][1]}")
        return matches_exactos_con_funda[0][1]
    
    # 2. Si no hay matches exactos, buscar por palabras clave con pesos
    print("\n🔎 No hay matches exactos, usando sistema de puntajes...")
    mejor_match = None
    mejor_puntaje = 0
    
    # Generar bigramas del texto sin medidas
    palabras_normalizadas = key_sin_medidas.split()
    bigramas = set()
    for i in range(len(palabras_normalizadas)-1):
        bigramas.add(f"{palabras_normalizadas[i]} {palabras_normalizadas[i+1]}")
    
    for p, info in MAPA_PRODUCTOS.items():
        puntaje = 0
        cat, modelo, color, talle = info
        
        # Normalizar también para comparación (sin medidas)
        p_sin_medidas = normalizar_texto_sin_medidas(p)
        modelo_norm = modelo.lower()
        color_norm = color.lower() if color else ""
        talle_norm = talle.lower() if talle else ""
        
        # Verificar si el modelo contiene el sufijo (solo funda)
        es_funda_modelo = "(solo funda)" in modelo.lower()
        
        # --- VERIFICACIÓN DE MODELO (MUCHO MÁS ESTRICTA) ---
        # Extraer la palabra principal del modelo (sin el sufijo de funda)
        modelo_para_verificar = modelo_norm.replace('(solo funda)', '').strip()
        palabras_modelo = modelo_para_verificar.split()
        modelo_principal = palabras_modelo[0] if palabras_modelo else ""
        
        # Verificar si la palabra principal del modelo está en el texto
        if modelo_principal and modelo_principal not in key_sin_medidas:
            # Si no está la palabra principal, descartar completamente
            continue
        
        # Verificar si hay al menos una palabra completa del modelo
        modelo_encontrado = False
        for palabra in palabras_modelo:
            if len(palabra) > 3 and palabra in key_sin_medidas:
                modelo_encontrado = True
                break
        
        if not modelo_encontrado and palabras_modelo:
            # Si ninguna palabra significativa del modelo aparece, descartar
            continue
        
        # Palabras clave del producto (sin medidas)
        palabras_clave = p_sin_medidas.split()
        
        # CONTAR COINCIDENCIAS BÁSICAS
        for palabra in palabras_clave:
            if len(palabra) > 2 and palabra in key_sin_medidas:
                puntaje += 2
        
        # BUSCAR BIGRAMAS
        palabras_p = p_sin_medidas.split()
        for i in range(len(palabras_p)-1):
            bigrama = f"{palabras_p[i]} {palabras_p[i+1]}"
            if bigrama in bigramas:
                puntaje += 5
        
        # PESOS ESPECIALES POR CATEGORÍA
        
        # ROPITA
        if cat == "ROPITA":
            if "remera" in key_sin_medidas or "buzo" in key_sin_medidas:
                puntaje += 5
            if color_norm in key_sin_medidas and ("remera" in key_sin_medidas or "buzo" in key_sin_medidas):
                puntaje += 10
            elif color_norm in key_sin_medidas and not ("remera" in key_sin_medidas or "buzo" in key_sin_medidas):
                puntaje -= 5
        
        # MANTA
        elif cat == "MANTA":
            if "manta" in key_sin_medidas or "mantita" in key_sin_medidas:
                puntaje += 5
            elif "doble faz" in key_sin_medidas:
                puntaje += 3
        
        # FUNDA - DETECCIÓN MEJORADA
        texto_habla_de_funda = ("funda" in key_sin_medidas or 
                               "solo funda" in key_sin_medidas or 
                               "repuesto" in key_sin_medidas)
        
        if texto_habla_de_funda:
            if es_funda_modelo:
                puntaje += 20
            else:
                puntaje -= 40
        else:
            if es_funda_modelo:
                puntaje -= 30
            elif "completa" in key_sin_medidas:
                puntaje += 5
        
        # BONUS POR MODELO COMPLETO
        if modelo_para_verificar in key_sin_medidas:
            puntaje += 10
        
        # BONUS POR COLOR
        if color_norm and color_norm in key_sin_medidas:
            puntaje += 4
        
        # ========== BONUS POR TALLE - MUCHO MÁS PESO ==========
        if talle_norm:
            talle_match = False
            
            # Buscar talle exacto como palabra independiente
            if re.search(r'\b' + talle_norm + r'\b', key_original):
                puntaje += 15
                talle_match = True
                print(f"    ✓ Talle exacto '{talle_norm}' encontrado: +15")
            
            # Buscar "talla X"
            elif re.search(r'talla\s*' + talle_norm, key_original):
                puntaje += 20
                talle_match = True
                print(f"    ✓ 'talla {talle_norm}' encontrado: +20")
            
            # Buscar talle entre paréntesis como (l)
            elif re.search(r'\(' + talle_norm + r'\)', key_original):
                puntaje += 18
                talle_match = True
                print(f"    ✓ '({talle_norm})' encontrado: +18")
            
            # Buscar talle con guión como "-l"
            elif re.search(r'-' + talle_norm + r'\b', key_original):
                puntaje += 16
                talle_match = True
                print(f"    ✓ '-{talle_norm}' encontrado: +16")
            
            # Penalizar si no hay match de talle y el texto tiene algún talle
            if not talle_match:
                # Buscar cualquier talle en el texto
                talles_en_texto = re.findall(r'\b(xs|s|m|l|xl|u)\b', key_original, re.IGNORECASE)
                if talles_en_texto:
                    # Penalización fuerte por talle incorrecto
                    puntaje -= 30
                    print(f"    ✗ Talle incorrecto: texto dice '{talles_en_texto[0]}', producto es '{talle_norm}' (-30)")
        
        # CASTIGO por palabras prohibidas
        for prohibida in palabras_prohibidas:
            if prohibida in key_sin_medidas and cat != "ROPITA":
                puntaje -= 10
        
        # TOLERANCIA A ESPACIOS EN "talla l"
        if "talla" in key_original and talle_norm:
            patron = f"talla\\s*{talle_norm}"
            if re.search(patron, key_original):
                puntaje += 3
        
        # CASTIGO si el texto es demasiado corto y no coincide con el modelo
        if len(key_sin_medidas.split()) < 3 and modelo_para_verificar not in key_sin_medidas:
            puntaje -= 20
        
        if puntaje > mejor_puntaje:
            mejor_puntaje = puntaje
            mejor_match = info
            print(f"    📊 Puntaje actual: {puntaje} ({info[1]} {info[3]} {info[2]})")
    
    umbral = 5
    if mejor_match and mejor_match[0] == "ROPITA":
        umbral = 8
    
    if mejor_puntaje >= umbral:
        print(f"  ✅ Match flexible (sin medidas): {mejor_match} (puntaje: {mejor_puntaje})")
        return mejor_match
    
    return None
# ══════════════════════════════════════════════════════════════════════════════
# Extaer los datos del envio
# ══════════════════════════════════════════════════════════════════════════════
def extraer_datos_envio(pdf_path):
    """
    Extrae de un PDF de pedidos los datos de envío por orden
    Retorna: dict con {numero_orden: {"fecha": str, "envio": str, "tipo": str}}
    """
    doc = fitz.open(pdf_path)
    texto_completo = ""
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        texto_completo += page.get_text() + "\n"
    
    doc.close()
    
    lineas = texto_completo.split('\n')
    datos_envio = {}
    i = 0
    
    print("📦 Extrayendo datos de envío...")
    
    while i < len(lineas):
        linea = lineas[i].strip()
        
        # Buscar orden
        if linea.startswith('Orden #'):
            match = re.search(r'Orden #(\d+)', linea)
            if match:
                num_orden = match.group(1)
                print(f"  Orden #{num_orden}")
                
                # Buscar fecha (línea siguiente suele tener "Realizada el ...")
                fecha = ""
                j = i + 1
                while j < len(lineas) and j < i + 10:
                    if "Realizada el" in lineas[j]:
                        fecha_match = re.search(r'Realizada el (\d{2}/\d{2}/\d{4})', lineas[j])
                        if fecha_match:
                            fecha = fecha_match.group(1)
                        break
                    j += 1
                
                # Buscar tipo de envío
                envio = "Desconocido"
                j = i + 1
                # Buscar en las siguientes 30 líneas
                while j < len(lineas) and j < i + 30:
                    linea_j = lineas[j].strip()
                    linea_j_lower = linea_j.lower()
                    
                    # Detectar por "Envío:"
                    if "envío:" in linea_j_lower or "envio:" in linea_j_lower:
                        if "andrea" in linea_j_lower:
                            envio = "Andreani"
                        break
                    
                    # Detectar por "Dirección de retiro:"
                    if "dirección de retiro:" in linea_j_lower or "direccion de retiro:" in linea_j_lower:
                        # Mirar las siguientes líneas para ver si es showroom o sucursal
                        k = j + 1
                        while k < len(lineas) and k < j + 10:
                            linea_k = lineas[k].strip().upper()
                            if "SHOWRROM" in linea_k or "SHOWROOM" in linea_k:
                                envio = "Retiro en Showroom"
                                break
                            elif "SUCURSAL ANDREANI" in linea_k:
                                envio = "Retiro Andreani"
                                break
                            k += 1
                        break
                    
                    j += 1
                
                datos_envio[num_orden] = {
                    "fecha": fecha if fecha else "No especificada",
                    "envio": envio,
                    "tipo_cliente": "MINORISTA"  # Todos minoristas
                }
                print(f"    Fecha: {fecha}")
                print(f"    Envío: {envio}")
        
        i += 1
    
    return datos_envio
# ══════════════════════════════════════════════════════════════════════════════
# PDF EXTRACTION (usando pymupdf)
# ══════════════════════════════════════════════════════════════════════════════
def extraer_ordenes_con_fitz(pdf_path):
    """
    Extrae órdenes y productos usando pymupdf
    """
    doc = fitz.open(pdf_path)
    texto_completo = ""
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        texto_completo += page.get_text() + "\n"
    
    doc.close()
    
    lineas = texto_completo.split('\n')
    ordenes = defaultdict(list)
    i = 0
    
    print("📄 Procesando páginas...")
    print(f"📋 Usando {len(PALABRAS_CLAVE)} palabras clave para detectar productos")
    
    while i < len(lineas):
        linea = lineas[i].strip()
        
        # Detectar inicio de orden
        if linea.startswith('Orden #'):
            match = re.search(r'Orden #(\d+)', linea)
            if match:
                num_orden = match.group(1)
                print(f"\n🔍 Procesando Orden #{num_orden}")
                
                # Recolectar SOLO las líneas de esta orden hasta la próxima "Orden #"
                lineas_orden = []
                j = i + 1
                while j < len(lineas):
                    siguiente_linea = lineas[j].strip()
                    # Si encontramos otra orden, terminamos
                    if siguiente_linea.startswith('Orden #'):
                        break
                    lineas_orden.append(lineas[j])
                    j += 1
                
                # Procesar SOLO las líneas de esta orden
                k = 0
                productos_en_orden = 0
                
                while k < len(lineas_orden):
                    linea_actual = lineas_orden[k].strip()
                    
                    # Detectar inicio de producto usando palabras clave dinámicas
                    if any(linea_actual.lower().startswith(p) for p in PALABRAS_CLAVE):
                        productos_en_orden += 1
                        print(f"\n  Producto #{productos_en_orden}:")
                        print(f"    L{k+1}: {linea_actual[:60]}...")
                        
                        # Capturar nombre del producto
                        nombre_producto = linea_actual
                        
                        m = k + 1
                        lineas_capturadas = 1
                        
                        print(f"    🔍 Buscando cantidad en líneas desde m={m}...")
                        
                        # Seguir capturando hasta encontrar la cantidad
                        cantidad_encontrada = False
                        while m < len(lineas_orden):
                            linea_m = lineas_orden[m].strip()
                            print(f"    L{m+1}: '{linea_m[:60]}'")
                            
                            # Si encontramos un número solo, es la cantidad
                            if linea_m.isdigit():
                                cantidad = int(linea_m)
                                print(f"    ✅ Cantidad encontrada: {cantidad}")
                                m += 1
                                cantidad_encontrada = True
                                break
                            
                            # Si encontramos "Subtotal", significa que no hay número antes
                            if "Subtotal" in linea_m:
                                cantidad = 1
                                print(f"    📌 Cantidad: {cantidad} (por defecto - encontró Subtotal)")
                                m += 1
                                cantidad_encontrada = True
                                break
                            
                            # Si encontramos "Cant." pero no es un número, seguir
                            if "Cant." in linea_m:
                                print(f"    📌 Línea 'Cant.' - continuando...")
                                m += 1
                                continue
                            
                            # 🔥 NUEVO: Si la línea empieza con "funda", siempre es parte del producto anterior
                            if linea_m.lower().startswith("funda"):
                                print(f"    🔄 Continuación (funda): {linea_m[:60]}...")
                                nombre_producto += " " + linea_m
                                lineas_capturadas += 1
                                m += 1
                                continue
                            
                            # Si encontramos otra palabra clave, es un NUEVO producto
                            if any(linea_m.lower().startswith(p) for p in PALABRAS_CLAVE):
                                print(f"    → Siguiente producto detectado (cantidad por defecto 1)")
                                cantidad = 1
                                cantidad_encontrada = True
                                break
                            
                            # Si no, es parte del mismo producto
                            nombre_producto += " " + linea_m
                            lineas_capturadas += 1
                            print(f"    L{lineas_capturadas}: {linea_m[:60]}...")
                            m += 1
                        
                        # Si salimos sin cantidad, establecer 1
                        if not cantidad_encontrada:
                            cantidad = 1
                            print(f"    ⚠️ No se encontró cantidad, usando 1")
                        
                        # Resolver producto
                        print(f"    📝 Nombre completo: {nombre_producto[:100]}...")
                        info = resolver(nombre_producto)
                        if info:
                            cat, modelo, color, talle = info
                            print(f"    ✅ → {modelo} {talle} {color} x{cantidad}")
                            ordenes[num_orden].append((info, cantidad))
                        else:
                            print(f"    ❌ ⚠ No resuelto: {nombre_producto[:80]}...")
                        
                        k = m
                    else:
                        k += 1
                
                print(f"  📊 Total productos en orden #{num_orden}: {productos_en_orden}")
                i = j
            else:
                i += 1
        else:
            i += 1
    
    # Agrupar productos por orden
    ordenes_agrupadas = {}
    for num_orden, productos in ordenes.items():
        print(f"\n📊 Orden #{num_orden} - productos sin agrupar: {len(productos)}")
        grupos = defaultdict(int)
        for info, cant in productos:
            grupos[info] += cant
        ordenes_agrupadas[num_orden] = [(info, cant) for info, cant in grupos.items()]
        print(f"   📦 Productos agrupados: {len(ordenes_agrupadas[num_orden])}")
        for info, cant in ordenes_agrupadas[num_orden]:
            print(f"     → {info[1]} {info[3]} {info[2]} x{cant}")
    
    return ordenes_agrupadas

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════
NC = 5
def fc(h): return PatternFill("solid",fgColor=h)
def bd(c='BBBBBB',s='thin'): x=Side(style=s,color=c); return Border(left=x,right=x,top=x,bottom=x)
C=Alignment(horizontal='center',vertical='center',wrap_text=True)
L=Alignment(horizontal='left',  vertical='center',wrap_text=True)

def build_excel(ordenes, catalogo, out_path, datos_envio=None):
    """
    Genera el Excel con catálogo, pedidos y resumen de envíos
    """
    print("\n📊 DEBUG - Productos recibidos para Excel:")
    print("-" * 50)
    
    det = defaultdict(int)
    
    for num_orden, productos in ordenes.items():
        print(f"\nOrden #{num_orden}:")
        for info, cant in productos:
            cat, modelo, color, talle = info
            print(f"  → {cat} | {modelo} | {color} | {talle} x{cant}")
            det[(cat, modelo, color, talle)] += cant
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo + Pedidos"
    
    # =========================================================
    # Hoja principal: Catálogo + Pedidos
    # =========================================================
    r = 1
    for sec in catalogo:
        cat = sec["cat"]
        hc, bc = CAT_COLORS.get(cat, ("607D8B", "ECEFF1"))
        
        # Título de categoría
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(row=r, column=1, value=cat)
        c.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        c.fill = fc(hc)
        c.alignment = C
        c.border = bd('888888', 'medium')
        for col in range(2, NC+1):
            ws.cell(row=r, column=col).border = bd('888888', 'medium')
        ws.row_dimensions[r].height = 22
        r += 1
        
        # Encabezados
        for i, h in enumerate(sec["headers"]):
            c = ws.cell(row=r, column=1+i, value=h)
            c.font = Font(name='Arial', bold=True, size=9, color='333333')
            c.fill = fc(bc)
            c.alignment = C
            c.border = bd()
        ws.row_dimensions[r].height = 30
        r += 1
        
        # Filas de productos
        for i, (modelo, color) in enumerate(sec["filas"]):
            bg = 'F9F9F9' if i % 2 else 'FFFFFF'
            cm = {}
            for ti, tk in enumerate(sec["talle_cols"]):
                v = det.get((cat, modelo, color, tk), 0)
                if v:
                    cm[2+ti] = v
            
            # Escribir modelo y color
            for j in range(NC):
                v = [modelo, color, '', '', ''][j]
                c = ws.cell(row=r, column=1+j, value=v)
                c.font = Font(name='Arial', size=9, bold=(j == 0))
                c.fill = fc(bg)
                c.alignment = L if j <= 1 else C
                c.border = bd()
            
            # Escribir cantidades
            for ci, val in cm.items():
                if val:
                    c = ws.cell(row=r, column=1+ci, value=val)
                    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
                    c.fill = fc(hc)
                    c.alignment = C
                    c.border = bd()
            
            ws.row_dimensions[r].height = 16
            r += 1
        
        ws.row_dimensions[r].height = 8
        r += 1
    
    # Ancho de columnas
    for col, w in zip('ABCDE', [22, 18, 17, 17, 17]):
        ws.column_dimensions[col].width = w
    
    # =========================================================
    # Hoja de Resumen Productos
    # =========================================================
    ws_resumen = wb.create_sheet("Resumen Productos")
    
    # Encabezados
    headers_resumen = ["Producto", "Modelo", "Color", "Talle", "Cantidad Total"]
    for i, header in enumerate(headers_resumen, 1):
        cell = ws_resumen.cell(row=1, column=i, value=header)
        cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        cell.fill = PatternFill("solid", fgColor="4A5568")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd()
    
    # Ancho de columnas
    ws_resumen.column_dimensions['A'].width = 40
    ws_resumen.column_dimensions['B'].width = 20
    ws_resumen.column_dimensions['C'].width = 20
    ws_resumen.column_dimensions['D'].width = 15
    ws_resumen.column_dimensions['E'].width = 15
    
    # Agrupar todos los productos de todas las órdenes
    resumen = defaultdict(int)
    productos_detalle = []
    
    for num_orden, productos in ordenes.items():
        for info, cant in productos:
            cat, modelo, color, talle = info
            clave = (modelo, color, talle)
            resumen[clave] += cant
            if clave not in [p[0] for p in productos_detalle]:
                nombre_producto = f"{modelo} {color} {talle}".strip()
                productos_detalle.append((clave, modelo, color, talle, nombre_producto))
    
    # Ordenar por cantidad
    productos_ordenados = sorted(
        [(clave, modelo, color, talle, nombre, resumen[clave]) 
         for (clave, modelo, color, talle, nombre) in productos_detalle],
        key=lambda x: -x[5]
    )
    
    # Escribir datos
    for row, (clave, modelo, color, talle, nombre, cantidad) in enumerate(productos_ordenados, 2):
        cell_a = ws_resumen.cell(row=row, column=1, value=nombre)
        cell_a.font = Font(name='Arial', size=10)
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        cell_a.border = bd()
        
        cell_b = ws_resumen.cell(row=row, column=2, value=modelo)
        cell_b.font = Font(name='Arial', size=10)
        cell_b.alignment = Alignment(horizontal='left', vertical='center')
        cell_b.border = bd()
        
        cell_c = ws_resumen.cell(row=row, column=3, value=color)
        cell_c.font = Font(name='Arial', size=10)
        cell_c.alignment = Alignment(horizontal='left', vertical='center')
        cell_c.border = bd()
        
        cell_d = ws_resumen.cell(row=row, column=4, value=talle)
        cell_d.font = Font(name='Arial', size=10)
        cell_d.alignment = Alignment(horizontal='center', vertical='center')
        cell_d.border = bd()
        
        cell_e = ws_resumen.cell(row=row, column=5, value=cantidad)
        cell_e.font = Font(name='Arial', bold=True, size=11)
        cell_e.fill = PatternFill("solid", fgColor="E9F0FA")
        cell_e.alignment = Alignment(horizontal='center', vertical='center')
        cell_e.border = bd()
    
    # Fila de total
    total_row = len(productos_ordenados) + 2
    cell_total = ws_resumen.cell(row=total_row, column=4, value="TOTAL:")
    cell_total.font = Font(name='Arial', bold=True, size=11)
    cell_total.alignment = Alignment(horizontal='right', vertical='center')
    cell_total.border = bd()
    
    cell_total_num = ws_resumen.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})")
    cell_total_num.font = Font(name='Arial', bold=True, size=11)
    cell_total_num.fill = PatternFill("solid", fgColor="E2E8F0")
    cell_total_num.alignment = Alignment(horizontal='center', vertical='center')
    cell_total_num.border = bd()
    
    if datos_envio:
        ws_envios = wb.create_sheet("Resumen Envíos")
        
        # ========== ENCABEZADO ==========
        # Título principal
        ws_envios.merge_cells('A1:D1')
        titulo = ws_envios.cell(row=1, column=1, value="SEGUIMIENTO COMANDAS")
        titulo.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
        titulo.fill = PatternFill("solid", fgColor="2C3E50")
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        titulo.border = bd()
        ws_envios.row_dimensions[1].height = 35
        
        # Fila de fecha y supervisor
        ws_envios.merge_cells('A2:B2')
        fecha_cell = ws_envios.cell(row=2, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        fecha_cell.font = Font(name='Arial', bold=True, size=11)
        fecha_cell.fill = PatternFill("solid", fgColor="E8F0F8")
        fecha_cell.alignment = Alignment(horizontal='left', vertical='center')
        fecha_cell.border = bd()
        
        ws_envios.merge_cells('C2:D2')
        supervisor_cell = ws_envios.cell(row=2, column=3, value="Supervisa: _________________________")
        supervisor_cell.font = Font(name='Arial', bold=True, size=11)
        supervisor_cell.fill = PatternFill("solid", fgColor="E8F0F8")
        supervisor_cell.alignment = Alignment(horizontal='left', vertical='center')
        supervisor_cell.border = bd()
        ws_envios.row_dimensions[2].height = 25
        
        # Espacio después del encabezado
        ws_envios.row_dimensions[3].height = 10
        
        # ========== TABLA DE DATOS ==========
        # Encabezados de tabla (fila 4)
        headers_envios = ["Comanda", "Fecha", "Envío", "Cliente Mayorista o MINORISTA"]
        header_fill = PatternFill("solid", fgColor="4A5568")
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        
        for i, header in enumerate(headers_envios, 1):
            cell = ws_envios.cell(row=4, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bd()
        ws_envios.row_dimensions[4].height = 28
        
        # Ancho de columnas
        ws_envios.column_dimensions['A'].width = 12
        ws_envios.column_dimensions['B'].width = 14
        ws_envios.column_dimensions['C'].width = 25
        ws_envios.column_dimensions['D'].width = 32
        
        # Ordenar números de comanda
        ordenes_ordenadas = sorted(datos_envio.keys(), key=lambda x: int(x) if x.isdigit() else 0)
        total_ordenes = len(ordenes_ordenadas)
        
        # Rango mínimo: hasta fila 43 (eso es 40 filas de datos desde la fila 5 a la 44)
        # fila 4 = encabezados, fila 5 = primera comanda
        min_fila_datos = 44  # Última fila del rango mínimo (5 a 44 son 40 filas)
        fila_actual = 5
        
        # Escribir todas las comandas
        for num_orden in ordenes_ordenadas:
            info = datos_envio[num_orden]
            
            # Alternar colores de fila
            bg_color = "F9F9F9" if fila_actual % 2 == 0 else "FFFFFF"
            
            # Comanda
            cell_a = ws_envios.cell(row=fila_actual, column=1, value=num_orden)
            cell_a.font = Font(name='Arial', size=10)
            cell_a.alignment = Alignment(horizontal='center', vertical='center')
            cell_a.fill = PatternFill("solid", fgColor=bg_color)
            cell_a.border = bd()
            
            # Fecha
            cell_b = ws_envios.cell(row=fila_actual, column=2, value=info["fecha"])
            cell_b.font = Font(name='Arial', size=10)
            cell_b.alignment = Alignment(horizontal='center', vertical='center')
            cell_b.fill = PatternFill("solid", fgColor=bg_color)
            cell_b.border = bd()
            
            # Envío
            cell_c = ws_envios.cell(row=fila_actual, column=3, value=info["envio"])
            cell_c.font = Font(name='Arial', size=10)
            cell_c.alignment = Alignment(horizontal='left', vertical='center')
            cell_c.fill = PatternFill("solid", fgColor=bg_color)
            cell_c.border = bd()
            
            # Tipo de cliente
            cell_d = ws_envios.cell(row=fila_actual, column=4, value=info["tipo_cliente"])
            cell_d.font = Font(name='Arial', size=10)
            cell_d.alignment = Alignment(horizontal='center', vertical='center')
            cell_d.fill = PatternFill("solid", fgColor=bg_color)
            cell_d.border = bd()
            
            fila_actual += 1
        
        # Si hay menos comandas que el mínimo, completar filas vacías hasta el mínimo
        fila_total_minima = max(fila_actual, min_fila_datos + 1)  # +1 porque fila_actual ya está en la siguiente
        
        for fila_vacia in range(fila_actual, fila_total_minima):
            for col in range(1, 5):
                cell = ws_envios.cell(row=fila_vacia, column=col, value="")
                cell.border = bd()
                if fila_vacia % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F9F9F9")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")
        
        # La fila del total va después de la última fila ocupada o del mínimo
        fila_total = max(fila_actual, min_fila_datos + 1)
        
        # Fila de total de órdenes
        ws_envios.merge_cells(f'A{fila_total}:C{fila_total}')
        cell_total = ws_envios.cell(row=fila_total, column=1, value=f"TOTAL DE COMANDAS: {total_ordenes}")
        cell_total.font = Font(name='Arial', bold=True, size=11)
        cell_total.alignment = Alignment(horizontal='right', vertical='center')
        cell_total.fill = PatternFill("solid", fgColor="E8F0F8")
        cell_total.border = bd()
        
        # Columna D del total queda vacía
        cell_d_total = ws_envios.cell(row=fila_total, column=4, value="")
        cell_d_total.border = bd()
        cell_d_total.fill = PatternFill("solid", fgColor="E8F0F8")
        
        ws_envios.row_dimensions[fila_total].height = 25
        
        print(f"✅ Hoja 'Resumen Envíos' creada con {total_ordenes} órdenes (filas hasta {fila_total})")
    
    # Guardar Excel
    wb.save(out_path)
    print(f"\n✅ Excel guardado en: {out_path}")
    
    return {}

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES PARA ANOTAR PDF DE ETIQUETAS
# ══════════════════════════════════════════════════════════════════════════════
def formatear_productos_orden(productos, resolver_func):
    """Agrupa productos iguales y devuelve líneas de texto en formato personalizado"""
    grupos = defaultdict(int)
    for info, cant in productos:
        cat, modelo, color, talle = info
        key = (cat, modelo, talle, color)
        grupos[key] += cant
    
    lineas = []
    for (cat, modelo, talle, color), cant in grupos.items():
        # Caso especial: MANTA
        if cat == "MANTA":
            color_simple = color.split('/')[0].lower()
            linea = f"Manta {color_simple} x{cant}"
        
        # Caso especial: GATITO
        elif modelo == "Gatito":
            if cat == "INVIERNO":
                linea = f"Gatito invierno {talle} x{cant}"
            elif cat == "VERANO":
                linea = f"Gatito verano {talle} x{cant}"
            else:
                linea = f"Gatito {talle} x{cant}"
        
        # Caso especial: HUELLA
        elif modelo == "Huella":
            if cat == "INVIERNO":
                linea = f"Huella invierno {talle} x{cant}"
            elif cat == "VERANO":
                linea = f"Huella verano {talle} x{cant}"
            elif cat == "ANTIESTRES":
                linea = f"Huella antiestrés {talle} x{cant}"
            else:
                linea = f"Huella {talle} x{cant}"
        
        # Caso especial: GARRA (solo cuando es realmente Garra)
        elif modelo == "Garra" and color == "Gris":
            if cat == "INVIERNO":
                linea = f"Garra invierno {talle} x{cant}"
            elif cat == "VERANO":
                linea = f"Garra verano {talle} x{cant}"
            elif cat == "ANTIESTRES":
                linea = f"Garra antiestrés {talle} x{cant}"
            else:
                linea = f"Garra {talle} x{cant}"
        
        # Caso especial: ROPITA
        elif cat == "ROPITA" and modelo == "Ropita":
            talle_num = {
                "XS": "N°1", "S": "N°2", "M": "N°3", "L": "N°4", "XL": "N°5", "U": ""
            }.get(talle, talle)
            
            if talle_num:
                linea = f"{color} {talle_num} x{cant}"
            else:
                linea = f"{color} x{cant}"
        
        else:
            if color and color.strip():
                linea = f"{modelo} {talle} {color} x{cant}"
            else:
                linea = f"{modelo} {talle} x{cant}"
        
        lineas.append(linea)
    
    return lineas

def anotar_pdf_con_productos(pdf_etiquetas_path, pdf_pedidos_paths, output_path):
    """
    Añade texto con productos justo debajo del último número de seguimiento
    Recibe una lista de paths de PDFs de pedidos
    """
    # Extraer órdenes de TODOS los PDFs de pedidos
    todas_ordenes = {}
    
    for pedido_path in pdf_pedidos_paths:
        ordenes = extraer_ordenes_con_fitz(pedido_path)
        # Combinar órdenes (si hay duplicados, se suman)
        for num_orden, productos in ordenes.items():
            if num_orden not in todas_ordenes:
                todas_ordenes[num_orden] = []
            todas_ordenes[num_orden].extend(productos)
    
    # Agrupar productos por orden (sumar cantidades)
    ordenes_agrupadas = {}
    for num_orden, productos in todas_ordenes.items():
        grupos = defaultdict(int)
        for info, cant in productos:
            grupos[info] += cant
        ordenes_agrupadas[num_orden] = [(info, cant) for info, cant in grupos.items()]
    
    # Abrir PDF de etiquetas
    doc = fitz.open(pdf_etiquetas_path)
    
    # 1 cm en puntos
    UN_CM = 28.35
    
    for pagina in doc:
        text = pagina.get_text()
        match = re.search(r"#(\d+)", text)
        if match:
            orden = match.group(1)
            if orden in ordenes_agrupadas:
                productos = ordenes_agrupadas[orden]
                lineas = formatear_productos_orden(productos, resolver)
                
                # Buscar el ÚLTIMO número de seguimiento
                palabras = pagina.get_text("words")
                seguimientos = [w for w in palabras if "seguimiento" in w[4].lower()]
                seguimientos.sort(key=lambda w: w[3])
                
                y_pos = None
                
                if len(seguimientos) >= 2:
                    segundo_seguimiento = seguimientos[1]
                    y_pos = segundo_seguimiento[3] + UN_CM
                elif len(seguimientos) == 1:
                    importantes = [w for w in palabras if "importante" in w[4].lower()]
                    if importantes:
                        importantes.sort(key=lambda w: w[3])
                        y_pos = importantes[-1][3] + UN_CM + 50
                
                if y_pos is None:
                    y_pos = pagina.rect.height - 80
                
                # Escribir texto
                tam_fuente = 16
                for i, linea in enumerate(lineas):
                    punto = fitz.Point(20, y_pos + (i * 18))
                    pagina.insert_text(punto, linea, fontsize=tam_fuente,
                                     fontname="helv", color=(0,0,0))
    
    doc.save(output_path)
    doc.close()
    return ordenes_agrupadas

def reorganizar_etiquetas(pdf_anotado_path, output_path, etiquetas_por_pagina=3):
    """
    Reorganiza un PDF anotado para poner múltiples etiquetas por página horizontal
    Con 0.5 cm de espacio al inicio de la PRIMER COMADA DE CADA HOJA
    """
    doc = fitz.open(pdf_anotado_path)
    output = fitz.open()
    
    # Configuración de página horizontal (A4 landscape)
    page_width_mm = 297
    page_height_mm = 210
    page_width_pt = page_width_mm * 2.83465
    page_height_pt = page_height_mm * 2.83465
    
    # Separación entre etiquetas
    spacing_mm = 5
    spacing_pt = spacing_mm * 2.83465
    
    # Margen superior para TODAS las etiquetas
    margin_top_mm = 30
    margin_top_pt = margin_top_mm * 2.83465
    
    # ESPACIO EXTRA PARA LA PRIMER COMADA DE CADA HOJA (0.5 cm)
    primer_comanda_extra_mm = 5  # 0.5 cm
    primer_comanda_extra_pt = primer_comanda_extra_mm * 2.83465
    
    total_paginas = len(doc)
    paginas_necesarias = (total_paginas + etiquetas_por_pagina - 1) // etiquetas_por_pagina
    
    print(f"\n📄 Reorganizando {total_paginas} etiquetas en {paginas_necesarias} páginas...")
    print(f"   📏 Cada primera comanda de cada hoja con +{primer_comanda_extra_mm} mm de margen izquierdo")
    
    for out_page_idx in range(paginas_necesarias):
        page = output.new_page(width=page_width_pt, height=page_height_pt)
        
        start_idx = out_page_idx * etiquetas_por_pagina
        end_idx = min(start_idx + etiquetas_por_pagina, total_paginas)
        
        current_x_pt = 0
        
        # EN CADA HOJA, la PRIMER COMADA tiene espacio extra
        current_x_pt += primer_comanda_extra_pt
        print(f"   📌 Página {out_page_idx + 1}: primera comanda con +{primer_comanda_extra_mm} mm de margen izquierdo")
        
        for i in range(start_idx, end_idx):
            src_page = doc[i]
            src_rect = src_page.rect
            
            # Obtener el contenido real
            text_dict = src_page.get_text("dict")
            blocks = text_dict.get("blocks", [])
            
            content_rect = None
            for block in blocks:
                if "bbox" in block:
                    bbox = fitz.Rect(block["bbox"])
                    if content_rect is None:
                        content_rect = bbox
                    else:
                        content_rect.include_rect(bbox)
            
            if content_rect is None or content_rect.is_empty:
                content_rect = src_rect
            
            # Escala 1:1
            scaled_width_pt = content_rect.width
            scaled_height_pt = content_rect.height
            
            target_rect = fitz.Rect(
                current_x_pt,
                margin_top_pt,
                current_x_pt + scaled_width_pt,
                margin_top_pt + scaled_height_pt
            )
            
            page.show_pdf_page(target_rect, doc, i, clip=content_rect)
            
            # Actualizar posición para la siguiente comanda
            current_x_pt += scaled_width_pt + spacing_pt
    
    output.save(output_path)
    output.close()
    doc.close()
    return output_path

# ══════════════════════════════════════════════════════════════════════════════
# FLASK ROUTES
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "pulguitas_ui.html")

@app.route("/productos", methods=["GET"])
def get_productos():
    txt = PRODUCTOS_TXT.read_text(encoding="utf-8") if PRODUCTOS_TXT.exists() else ""
    return jsonify({"content": txt})

@app.route("/productos", methods=["POST"])
def save_productos():
    data = request.get_json()
    PRODUCTOS_TXT.write_text(data["content"], encoding="utf-8")
    return jsonify({"ok": True})

@app.route("/mapeo", methods=["GET"])
def get_mapeo():
    """Devuelve todos los productos desde la BD en formato JSON anidado"""
    datos = exportar_mapeo_a_json()
    return jsonify({"contenido": json.dumps(datos, ensure_ascii=False)})

@app.route("/mapeo", methods=["POST"])
def save_mapeo():
    """Guarda productos en la BD"""
    data = request.get_json()
    contenido = data.get("contenido", "{}")
    
    if guardar_productos_desde_json(contenido):
        global MAPA_PRODUCTOS, CONFIG, PALABRAS_CLAVE
        MAPA_PRODUCTOS = cargar_mapeo_desde_bd()
        CONFIG = cargar_configuracion_desde_mapeo()
        PALABRAS_CLAVE = generar_palabras_clave()
        return jsonify({"ok": True})
    else:
        return jsonify({"error": "Error guardando en BD"}), 500

# ── Rutas para categorías ───────────────────────────────────────────────────
@app.route("/categorias", methods=["GET"])
def get_categorias():
    """Obtiene todas las categorías"""
    categorias = get_all_categorias()
    return jsonify([{
        "nombre": c.nombre,
        "color_fondo": c.color_fondo,
        "color_texto": c.color_texto
    } for c in categorias])

@app.route("/categorias", methods=["POST"])
def crear_categoria():
    """Crea una nueva categoría"""
    data = request.get_json()
    nombre = data.get("nombre", "").upper()
    color_fondo = data.get("color_fondo", "607D8B")
    color_texto = data.get("color_texto", "FFFFFF")
    
    if not nombre:
        return jsonify({"error": "Nombre requerido"}), 400
    
    if guardar_categoria(nombre, color_fondo, color_texto):
        global CAT_COLORS
        CAT_COLORS = get_colores_categorias()  # Recargar colores
        return jsonify({"ok": True})
    return jsonify({"error": "Error creando categoría"}), 500

@app.route("/categorias/<nombre>", methods=["DELETE"])
def eliminar_categoria_route(nombre):
    """Elimina una categoría"""
    if eliminar_categoria(nombre):
        global CAT_COLORS
        CAT_COLORS = get_colores_categorias()  # Recargar colores
        return jsonify({"ok": True})
    return jsonify({"error": "Error eliminando categoría"}), 500

@app.route("/categorias/<nombre>/colores", methods=["PUT"])
def actualizar_colores(nombre):
    """Actualiza los colores de una categoría"""
    data = request.get_json()
    color_fondo = data.get("color_fondo", "607D8B")
    color_texto = data.get("color_texto", "FFFFFF")
    
    if actualizar_colores_categoria(nombre, color_fondo, color_texto):
        global CAT_COLORS
        CAT_COLORS = get_colores_categorias()  # Recargar colores
        return jsonify({"ok": True})
    return jsonify({"error": "Error actualizando colores"}), 500

@app.route("/analizar", methods=["POST"])
def analizar():
    # Verificar que se recibieron archivos
    if "pdf" not in request.files:
        return jsonify({"error": "No se recibió PDF"}), 400
    
    pdf_files = request.files.getlist("pdf")
    if len(pdf_files) == 0:
        return jsonify({"error": "No se seleccionó ningún archivo"}), 400
    
    prod_txt = request.form.get("productos", "")
    
    # Crear archivos temporales para cada PDF
    temp_paths = []
    try:
        for pdf_file in pdf_files:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                pdf_file.save(tmp.name)
                temp_paths.append(tmp.name)
        
        # Combinar todos los pedidos de todos los PDFs
        todas_ordenes = {}
        todos_datos_envio = {}
        ordenes_sin_productos = set()  # Conjunto para órdenes sin productos
        
        for tmp_path in temp_paths:
            ordenes = extraer_ordenes_con_fitz(tmp_path)
            datos_envio = extraer_datos_envio(tmp_path)
            
            # Combinar órdenes
            for num_orden, productos in ordenes.items():
                if num_orden not in todas_ordenes:
                    todas_ordenes[num_orden] = []
                todas_ordenes[num_orden].extend(productos)
            
            # Combinar datos de envío
            todos_datos_envio.update(datos_envio)
        
        # Identificar órdenes sin productos (existen en datos_envio pero no en ordenes)
        for num_orden in todos_datos_envio.keys():
            if num_orden not in todas_ordenes or len(todas_ordenes[num_orden]) == 0:
                ordenes_sin_productos.add(num_orden)
        
        # Agrupar productos por orden (sumar cantidades)
        ordenes_agrupadas = {}
        for num_orden, productos in todas_ordenes.items():
            grupos = defaultdict(int)
            for info, cant in productos:
                grupos[info] += cant
            ordenes_agrupadas[num_orden] = [(info, cant) for info, cant in grupos.items()]
        
        # Generar catálogo y Excel
        catalogo = cargar_catalogo(prod_txt)
        sin_mapear = build_excel(ordenes_agrupadas, catalogo, str(OUTPUT_XLSX), todos_datos_envio)
        
        rows = []
        total_unidades = 0
        for num_orden, productos in ordenes_agrupadas.items():
            for info, cant in productos:
                cat, modelo, color, talle = info
                total_unidades += cant
                rows.append({
                    "producto": f"Orden #{num_orden}",
                    "cantidad": cant,
                    "categoria": cat,
                    "modelo": modelo,
                    "color": color,
                    "talle": talle,
                })
        
        return jsonify({
            "ok": True,
            "total_tipos": len(ordenes_agrupadas),
            "total_unidades": total_unidades,
            "sin_mapear": len(ordenes_sin_productos),
            "sin_mapear_list": [],
            "ordenes_sin_productos": sorted(list(ordenes_sin_productos), key=lambda x: int(x) if x.isdigit() else 0),
            "rows": rows,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        for tmp_path in temp_paths:
            try:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
            except:
                pass

@app.route("/anotar", methods=["POST"])
def anotar():
    """Recibe múltiples PDFs de pedidos y un PDF de etiquetas, anota y reorganiza automáticamente"""
    if "pedidos" not in request.files or "etiquetas" not in request.files:
        return jsonify({"error": "Se requieren dos archivos: pedidos PDF y etiquetas PDF"}), 400
    
    pedidos_files = request.files.getlist("pedidos")  # Obtener TODOS los archivos de pedidos
    etiquetas_file = request.files["etiquetas"]
    
    tmp_pedidos_paths = []
    for pedidos_file in pedidos_files:
        with tempfile.NamedTemporaryFile(suffix="_pedidos.pdf", delete=False) as tmp_pedidos:
            pedidos_file.save(tmp_pedidos.name)
            tmp_pedidos_paths.append(tmp_pedidos.name)
    
    with tempfile.NamedTemporaryFile(suffix="_etiquetas.pdf", delete=False) as tmp_etiquetas:
        etiquetas_file.save(tmp_etiquetas.name)
        tmp_etiquetas_path = tmp_etiquetas.name
    
    anotado_path = tmp_etiquetas_path.replace("_etiquetas", "_anotado")
    final_path = tmp_etiquetas_path.replace("_etiquetas", "_final")
    
    try:
        print(f"📝 PASO 1: Anotando PDF con productos (usando {len(tmp_pedidos_paths)} archivos de pedidos)...")
        anotar_pdf_con_productos(tmp_etiquetas_path, tmp_pedidos_paths, anotado_path)
        
        print("📐 PASO 2: Reorganizando PDF (3 etiquetas por página)...")
        reorganizar_etiquetas(anotado_path, final_path, etiquetas_por_pagina=3)
        
        print("✅ Proceso completado. Enviando PDF final...")
        return send_file(final_path, as_attachment=True, 
                        download_name=f"final_{etiquetas_file.filename}",
                        mimetype="application/pdf")
    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        for p in tmp_pedidos_paths:
            try:
                if os.path.exists(p):
                    os.unlink(p)
            except:
                pass
        for p in [tmp_etiquetas_path, anotado_path, final_path]:
            try:
                if os.path.exists(p):
                    os.unlink(p)
            except:
                pass

@app.route("/descargar")
def descargar():
    if not OUTPUT_XLSX.exists():
        return "No hay Excel generado aún", 404
    return send_file(str(OUTPUT_XLSX), as_attachment=True,
                     download_name="resumen_pedidos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/admin_productos.html")
def admin_productos():
    return send_from_directory(BASE_DIR, "admin_productos.html")

if __name__ == "__main__":
    port = 5173
    url  = f"http://localhost:{port}"
    print(f"\n🐾 Pulguitas App corriendo en {url}")
    print("   Ctrl+C para detener\n")
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    app.run(port=port, debug=False)
